from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from .models import Student, AttendanceRecord, TimeTable, TeacherSubject, Notification, AssessmentRequest, AccessoryRequest, TeacherProfile, StoreStaff, StoreRequest, StoreRequestItem, StoreNotification, CourseMaterial, StudentSubmission, LateSubmissionRequest, ClassCoordinator, StudentApplication
from .utils import train_model, identify_faces, detect_and_crop_face, get_existing_attendance_record
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from django.db import transaction
from django.db.models import Q, Count
import os
import datetime
import json
import base64
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required
from django.utils import timezone

def login_view(request):
    if request.user.is_authenticated:
        return redirect('index')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user.is_superuser:
                return redirect('admin_dashboard')
            # Store Staff / Store Head redirect
            try:
                sp = user.store_profile
                if sp.role == 'head':
                    return redirect('store_head_dashboard')
                else:
                    return redirect('store_staff_tasks')
            except Exception:
                pass
            # Teacher redirect
            if user.is_staff:
                return redirect('teacher_dashboard')
            # Student redirect
            try:
                student = user.student
                if not student.is_registered:
                    logout(request)
                    messages.warning(request, "Your account has not been activated. Please verify details and register your face to activate your account.")
                    return redirect('register')
            except Student.DoesNotExist:
                pass
            return redirect('student_dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def register(request):
    return render(request, 'register.html')

@csrf_exempt
def register_verify(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            roll_number = data.get('roll_number')
            password = data.get('password')
        except Exception:
            return JsonResponse({'status': 'error', 'message': 'Invalid request data.'})
            
        user = authenticate(request, username=roll_number, password=password)
        if user is not None:
            try:
                student = user.student
                if student.is_registered:
                    return JsonResponse({'status': 'error', 'message': 'Account already activated. Please login directly.'})
                return JsonResponse({
                    'status': 'ok',
                    'student': {
                        'name': student.name,
                        'roll_number': student.roll_number,
                        'department': student.department or '',
                        'year': student.year or '',
                        'section': student.section or ''
                    }
                })
            except Student.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'No student profile found linked to this Roll Number.'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid Roll Number or Password.'})
            
    return JsonResponse({'status': 'error', 'message': 'GET method not supported.'})

@csrf_exempt
def register_activate(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            roll_number = data.get('roll_number')
            password = data.get('password')
            new_password = data.get('new_password')
            images = data.get('images')
        except Exception:
            return JsonResponse({'status': 'error', 'message': 'Invalid request data.'})
            
        user = authenticate(request, username=roll_number, password=password)
        if user is None:
            return JsonResponse({'status': 'error', 'message': 'Authentication failed.'})
            
        try:
            student = user.student
            if student.is_registered:
                return JsonResponse({'status': 'error', 'message': 'Account already activated.'})
        except Student.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student profile not found.'})
            
        if not images or len(images) < 5:
            return JsonResponse({'status': 'error', 'message': 'Please capture all 5 required face angles.'})
            
        user.set_password(new_password)
        user.save()
        student.plain_password = new_password
        student.save()
        
        folder_name = f"{student.roll_number}_{student.name}"
        safe_dept = "".join([c for c in (student.department or 'General') if c.isalnum() or c in (' ', '_', '-')]).strip()
        safe_year = "".join([c for c in str(student.year or '1') if c.isalnum()]).strip()
        safe_section = "".join([c for c in str(student.section or 'A') if c.isalnum()]).strip()
        
        save_dir = os.path.join(settings.DATASET_DIR, safe_dept, safe_year, safe_section, folder_name)
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
            
        fs = FileSystemStorage()
        success_count = 0
        angle_names = ["center", "left", "right", "up", "down"]
        
        for idx, base64_str in enumerate(images):
            try:
                if ';base64,' in base64_str:
                    img_data = base64_str.split(';base64,')[1]
                else:
                    img_data = base64_str
                    
                decoded_img = base64.b64decode(img_data)
                
                filename = fs.save(f"temp_reg_{student.roll_number}_{angle_names[idx]}.jpg", ContentFile(decoded_img))
                temp_path = fs.path(filename)
                
                if detect_and_crop_face(temp_path, save_dir, folder_name):
                    success_count += 1
                    
                fs.delete(filename)
            except Exception as e:
                print(f"Error processing image {idx}: {e}")
                
        if success_count == 0:
            return JsonResponse({'status': 'error', 'message': 'No face could be detected in any of the captured images. Please capture again in a well-lit room.'})
            
        train_success, train_msg = train_model()
        if not train_success:
            print(f"Model training failed: {train_msg}")
            
        student.is_registered = True
        student.save()
        
        # Log user in
        login(request, user)
        
        return JsonResponse({
            'status': 'ok',
            'message': f'Account activated successfully! {success_count} face photos processed and model trained.'
        })

def index(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        elif request.user.is_staff:
            return redirect('teacher_dashboard')
        else:
            return redirect('student_dashboard')
    return redirect('login')

def home(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        elif request.user.is_staff:
            return redirect('teacher_dashboard')
        else:
            try:
                sp = request.user.store_profile
                if sp.role == 'head':
                    return redirect('store_head_dashboard')
                else:
                    return redirect('store_staff_tasks')
            except Exception:
                pass
            return redirect('student_dashboard')
            
    total_students = Student.objects.count()
    total_attendance = AttendanceRecord.objects.count()
    return render(request, 'home.html', {'total_students': total_students, 'total_attendance': total_attendance})



def train(request):
    if request.method == 'POST':
        success, msg = train_model()
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        elif request.user.is_staff:
            return redirect('teacher_dashboard')
        else:
            return redirect('student_dashboard')
    return redirect('login')

@login_required
def upload_attendance(request):
    # Get params
    subject = request.GET.get('subject') or request.POST.get('subject')
    year = request.GET.get('year') or request.POST.get('year')
    section = request.GET.get('section') or request.POST.get('section')
    
    # 1. Security Check: Ensure Teacher is assigned to this Subject/Year
    if not request.user.is_superuser:
        if not TeacherSubject.objects.filter(teacher=request.user, subject=subject, year=year).exists():
             messages.error(request, "You are not authorized to mark attendance for this class.")
             return redirect('teacher_dashboard')

    if request.method == 'POST' and request.FILES.get('class_image'):
        image = request.FILES['class_image']
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads'), base_url='/media/uploads/')
        filename = fs.save(image.name, image)
        file_path = fs.path(filename)
        
        predictions = identify_faces(file_path)
        
        marked_count = 0
        unknown_count = 0
        absent_count = 0
        
        checked_rolls = set()
        present_roll_numbers = set()
        final_predictions = []
        
        today = datetime.date.today()
        
        # 1. Process Detected Faces (Mark Present)
        for pred in predictions:
            roll_number = pred['roll_number']
            if roll_number:
                present_roll_numbers.add(roll_number)
                
                # Deduplication: Check if we already processed this student in this frame
                if roll_number in checked_rolls:
                    continue
                
                checked_rolls.add(roll_number)
                
                try:
                    student = Student.objects.get(roll_number=roll_number)
                    
                    # Verify student belongs to this year/section
                    if year and str(student.year) != str(year):
                        pred['status'] = f"Wrong Year ({student.year})"
                        final_predictions.append(pred)
                        continue
                    if section and str(student.section).lower() != str(section).lower():
                        pred['status'] = f"Wrong Section ({student.section})"
                        final_predictions.append(pred)
                        continue
                    # Verify student belongs to this department/branch if section is a department code
                    dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
                    if section and section.upper() in dept_codes:
                        if not student.department or section.lower() not in student.department.lower():
                            pred['status'] = f"Wrong Branch ({student.department})"
                            final_predictions.append(pred)
                            continue
                        
                    status_text = "Marked Present"
                    if subject:
                         status_text += f" ({subject})"
                    
                    can_mark = True
                    
                    if subject:
                         # Use unified helper for duplicate check
                         # Fix: Use local time for check matching TimeTable slots
                         existing_record = get_existing_attendance_record(student, subject, today, datetime.datetime.now().time())
                         
                         if existing_record:
                             if existing_record.status == 'Absent':
                                 existing_record.status = 'Present'
                                 existing_record.save()
                                 status_text = f"Marked Present (Was Absent) - ({subject})"
                                 marked_count += 1
                                 can_mark = False # Handled update
                             else:
                                 can_mark = False
                                 status_text = f"Already Marked Present ({subject})"
                    else:
                        # Fallback Global Cooldown
                        last_attendance = AttendanceRecord.objects.filter(student=student).order_by('-date', '-time').first()
                        if last_attendance:
                            last_datetime_naive = datetime.datetime.combine(last_attendance.date, last_attendance.time)
                            if timezone.is_naive(last_datetime_naive):
                                last_datetime = timezone.make_aware(last_datetime_naive, timezone.get_current_timezone())
                            else:
                                last_datetime = last_datetime_naive
                                
                            time_diff = timezone.now() - last_datetime
                            if time_diff.total_seconds() < 3600 and last_attendance.status == 'Present':
                                can_mark = False
                                status_text = f"Already Marked ({int(time_diff.total_seconds() // 60)}m ago)"

                    if can_mark:
                         AttendanceRecord.objects.create(student=student, subject=subject, status='Present')
                         marked_count += 1
                         
                    pred['status'] = status_text
                    
                except Student.DoesNotExist:
                    print(f"DEBUG: Student with roll number {roll_number} not found in DB.")
                    pred['status'] = 'Student Not Found'
                    pass
            else:
                unknown_count += 1
                pred['status'] = 'Unknown'
            
            final_predictions.append(pred)

        # 2. Process Missing Students (Auto-Absent) - ONLY if Subject is defined (Class Mode)
        # CRITICAL FIX: Only fetch students for this specific YEAR (and section)
        if subject and year:
            filter_kwargs = {'year': year}
            if section:
                filter_kwargs['section'] = section
                
            class_students = Student.objects.filter(**filter_kwargs)
            # If section name is a department code, ensure student's department also matches to avoid cross-branch matching
            dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
            if section and section.upper() in dept_codes:
                class_students = class_students.filter(department__icontains=section)
            
            for student in class_students:
                if student.roll_number not in present_roll_numbers:
                    # Check if recent record exists
                    recent_exists = False
                    cutoff_time = timezone.now() - datetime.timedelta(minutes=60)
                    recs = AttendanceRecord.objects.filter(student=student, date=today, subject=subject)
                    for r in recs:
                        dt_n = datetime.datetime.combine(r.date, r.time)
                        if timezone.is_naive(dt_n):
                            dt_a = timezone.make_aware(dt_n, timezone.get_current_timezone())
                        else:
                            dt_a = dt_n
                        if dt_a > cutoff_time:
                            recent_exists = True
                            break
                            
                    if not recent_exists:
                         AttendanceRecord.objects.create(student=student, subject=subject, status='Absent')
                         absent_count += 1
                
        messages.success(request, f"Present: {marked_count}, Absent: {absent_count}, Unknown: {unknown_count}")
        return render(request, 'attendance_result.html', {'predictions': final_predictions, 'image_url': fs.url(filename)})
        
    return render(request, 'upload_attendance.html', {'subject': subject, 'year': year, 'section': section})

def attendance_list(request):
    records = AttendanceRecord.objects.all().order_by('-date', '-time')
    return render(request, 'attendance_list.html', {'records': records})

def student_attendance(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    records = AttendanceRecord.objects.filter(student=student).order_by('-date')
    return render(request, 'student_attendance.html', {'student': student, 'records': records})

@login_required
def manual_attendance(request):
    subject_raw = request.GET.get('subject')
    subject = subject_raw.strip() if subject_raw else ''
    year = request.GET.get('year')
    section = request.GET.get('section')

    # Fetch students for this class
    students = Student.objects.none()
    if year:
        students = Student.objects.filter(year__icontains=year)
        if section:
            students = students.filter(section__icontains=section)
            # If section name is a department code, ensure student's department also matches to avoid cross-branch matching
            dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
            if section.upper() in dept_codes:
                students = students.filter(department__icontains=section)
        students = students.order_by('roll_number')

    if request.method == 'POST':
        date_str = request.POST.get('date')
        try:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect(request.path)
            
        # Get list of student IDs marked as present
        present_student_ids = request.POST.getlist('student_ids')
        
        count_present = 0
        count_absent = 0
        
        time_str = request.POST.get('time')
        current_time = datetime.datetime.now().time()
        
        if time_str:
            try:
                current_time = datetime.datetime.strptime(time_str, '%H:%M').time()
            except ValueError:
                pass # Fallback to now
        
        with transaction.atomic():
            for student in students:
                # Lock the student row to prevent race conditions (double submission)
                # We simply fetch it with lock; we can continue using the loop variable 'student' or the locked instance.
                # Since 'student' variable is already populated, we just hold the lock.
                Student.objects.select_for_update().get(id=student.id)
                
                is_present = str(student.id) in present_student_ids
                status = 'Present' if is_present else 'Absent'
            
                # Unified lookup
                # Use provided time or current time for reference
                ref_time = current_time
                record = get_existing_attendance_record(student, subject, date_obj, ref_time)

                if record:
                    # Update existing
                    record.status = status
                    # Update time if beneficial, or keep original?
                    # If we are "correcting" a record, we might keep time. If we are explicitly setting time, we set it.
                    if time_str:
                         record.time = current_time
                    record.save()
                    create_notification(student, f"Attendance updated: {status} for {subject} on {date_obj} at {current_time.strftime('%H:%M')}")
                else:
                    # Create new
                    AttendanceRecord.objects.create(
                        student=student,
                        date=date_obj,
                        subject=subject,
                        status=status,
                        time=current_time
                    )
                    create_notification(student, f"Attendance marked: {status} for {subject} on {date_obj} at {current_time.strftime('%H:%M')}")
            
                if is_present:
                    count_present += 1
                else:
                    count_absent += 1
            
        messages.success(request, f"Attendance marked for {subject}: {count_present} Present, {count_absent} Absent.")
        return redirect('teacher_dashboard')


    today = datetime.date.today().strftime('%Y-%m-%d')
    default_time = request.GET.get('start_time', datetime.datetime.now().strftime('%H:%M'))
    
    # Check if attendance exists for this slot
    present_student_ids = []
    attendance_marked = False
    
    if default_time:
         try:
            query_time = datetime.datetime.strptime(default_time, '%H:%M').time()
            cutoff_start = datetime.datetime.combine(datetime.date.today(), query_time) - datetime.timedelta(minutes=20)
            cutoff_end = datetime.datetime.combine(datetime.date.today(), query_time) + datetime.timedelta(minutes=20)
            
            if timezone.is_aware(timezone.now()):
                 cutoff_start = timezone.make_aware(cutoff_start)
                 cutoff_end = timezone.make_aware(cutoff_end)
            
            existing_records = AttendanceRecord.objects.filter(
                subject=subject, 
                date=datetime.date.today()
            )
            
            # Filter in python or complex query for time range
            # Since we removed auto_now_add, we trust the time field more.
            # Let's find records in the window.
            matched_records = []
            for r in existing_records:
                dt_naive = datetime.datetime.combine(r.date, r.time)
                if timezone.is_aware(timezone.now()):
                    dt_check = timezone.make_aware(dt_naive, timezone.get_current_timezone())
                else:
                    dt_check = dt_naive
                
                if cutoff_start <= dt_check <= cutoff_end:
                    matched_records.append(r)
            
            if matched_records:
                attendance_marked = True
                for r in matched_records:
                    if r.status == 'Present':
                        present_student_ids.append(r.student.id)

         except ValueError:
             pass

    return render(request, 'manual_attendance.html', {
        'students': students, 
        'subject': subject, 
        'year': year, 
        'section': section,
        'today': today,
        'default_time': default_time,
        'attendance_marked': attendance_marked,
        'present_student_ids': present_student_ids,
    })

def live_attendance(request):
    subject = request.GET.get('subject')
    year = request.GET.get('year')
    section = request.GET.get('section')
    return render(request, 'live_attendance.html', {
        'subject': subject,
        'year': year,
        'section': section
    })

@csrf_exempt
def process_live_frame(request):
    import cv2
    import numpy as np
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            image_data = data.get('image')
            
            # Get explicit class params
            req_subject = data.get('subject')
            req_year = data.get('year')
            req_section = data.get('section')
            
            print(f"Received live frame request. params: {req_subject}, {req_year}, {req_section}")
            
            if not image_data:
                print("No image data received")
                return JsonResponse({'status': 'error', 'message': 'No image data'})

            # Decode base64
            if ';base64,' in image_data:
                format, imgstr = image_data.split(';base64,') 
            else:
                imgstr = image_data

            try:
                img_bytes = base64.b64decode(imgstr)
            except Exception as e:
                print(f"Base64 decode error: {e}")
                return JsonResponse({'status': 'error', 'message': 'Invalid base64 data'})
            
            # Convert to numpy array
            nparr = np.frombuffer(img_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if img is None:
                 print("Failed to decode image from numpy array")
                 return JsonResponse({'status': 'error', 'message': 'Failed to decode image'})

            # Convert BGR to RGB (face_recognition uses RGB)
            rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            
            print("Calling identify_faces...")
            predictions = identify_faces(image_content=rgb_img)
            print(f"Predictions: {predictions}")
            
            if req_subject:
                # Explicit mode
                current_subject = req_subject
            else:
                # TimeTable mode
                # Get current subject from timetable
                now = datetime.datetime.now()
                current_time = now.time()
                current_weekday = now.weekday()
                
                # Find active class
                # We filter for classes that started before now and end after now
                active_class = TimeTable.objects.filter(
                    day=current_weekday,
                    start_time__lte=current_time,
                    end_time__gte=current_time
                ).first()
                
                current_subject = active_class.subject if active_class else "Extra Class"

            # Mark attendance
            results = []
            
            # Deduplication for this frame
            req_time_str = data.get('start_time')
            
            # Default to now if not provided
            current_time = datetime.datetime.now().time()
            if req_time_str:
                try:
                    current_time = datetime.datetime.strptime(req_time_str, '%H:%M').time()
                except ValueError:
                    pass

            # ... decode image ... (omitted in tool, matching context below)

            # Mark attendance logic
            
            # Deduplication for this frame
            processed_rolls_in_frame = set()
            
            for pred in predictions:
                status_msg = ""
                roll_number = pred['roll_number']
                
                if roll_number:
                     # Check if we already processed this student in this frame
                    if roll_number in processed_rolls_in_frame:
                        continue # Skip duplicate
                    else:
                        processed_rolls_in_frame.add(roll_number)
                        try:
                            # ATOMIC BLOCK START: Lock student to prevent race condition duplicates
                            with transaction.atomic():
                                student = Student.objects.select_for_update().get(roll_number=roll_number)
                                
                                # Validate logic if year/section provided
                                if req_year:
                                    if str(student.year) != str(req_year):
                                        continue # Skip wrong year
                                if req_section:
                                    if str(student.section).lower() != str(req_section).lower():
                                        continue # Skip wrong section
                                    
                                    # Verify student belongs to this department/branch if section is a department code
                                    dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
                                    if req_section.upper() in dept_codes:
                                        if not student.department or req_section.lower() not in student.department.lower():
                                            continue # Skip wrong branch/department
                                        
                                can_mark = True
                                
                                if req_subject:
                                     # Explicit Subject Mode (Dashboard)
                                     # Use unified helper
                                     existing_record = get_existing_attendance_record(student, req_subject, datetime.date.today(), current_time)
                                     
                                     if existing_record:
                                         can_mark = False
                                         status_msg = f"Already Marked ({existing_record.time.strftime('%H:%M')})"
                                else:
                                    # Fallback / Timetable Mode (Auto-detect)
                                    # Use Global Cooldown (1 hour) to prev spam
                                    last_attendance = AttendanceRecord.objects.filter(student=student).order_by('-date', '-time').first()
                                    if last_attendance:
                                        last_datetime_naive = datetime.datetime.combine(last_attendance.date, last_attendance.time)
                                        if timezone.is_aware(timezone.now()):
                                             last_datetime = timezone.make_aware(last_datetime_naive, datetime.timezone.utc)
                                        else:
                                             last_datetime = last_datetime_naive
                                        
                                        time_diff = timezone.now() - last_datetime
                                        if time_diff.total_seconds() < 3600:
                                            can_mark = False
                                            status_msg = f"Already Marked ({int(time_diff.total_seconds() // 60)}m ago)"
                                
                                if can_mark:
                                    try:
                                        AttendanceRecord.objects.create(student=student, subject=current_subject, time=current_time)
                                        create_notification(student, f"Marked Present for {current_subject} via Face ID")
                                        status_msg = f"Marked Present ({current_subject})"
                                    except Exception as e:
                                         status_msg = f"Already Marked Today ({current_subject})"
                                    
                        except Student.DoesNotExist:
                            status_msg = "Student Not Found"
                else:
                    status_msg = "Unknown"
                
                results.append({
                    'name': pred['name'],
                    'roll_number': roll_number,
                    'location': pred['location'],
                    'status': status_msg
                })
            
            return JsonResponse({'status': 'success', 'results': results})
            
        except Exception as e:
            print(f"Error in process_live_frame: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@user_passes_test(lambda u: u.is_superuser)
def add_teacher(request):
    from .models import TeacherProfile
    dept_choices = TeacherProfile.DEPARTMENT_CHOICES
    if request.method == 'POST':
        username    = request.POST.get('username')
        email       = request.POST.get('email')
        password    = request.POST.get('password')
        designation = request.POST.get('designation', 'asst_prof')
        department  = request.POST.get('department', 'CSE')

        # Coordinator fields
        is_coord    = request.POST.get('is_coordinator') == 'on'
        coord_dept  = request.POST.get('coord_department', '').strip()
        coord_year  = request.POST.get('coord_year', '').strip()
        coord_sec   = request.POST.get('coord_section', '').strip()

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return render(request, 'add_teacher.html', {'dept_choices': dept_choices})

        if is_coord:
            if not coord_dept or not coord_year or not coord_sec:
                messages.error(request, "All coordinator class fields (Department, Year, Section) are required when 'Assign as Class Coordinator' is checked.")
                return render(request, 'add_teacher.html', {'dept_choices': dept_choices})

            existing_other = ClassCoordinator.objects.filter(
                department=coord_dept,
                year=coord_year,
                section=coord_sec
            ).first()
            if existing_other:
                messages.error(request, f"Class {coord_dept} Year {coord_year}-{coord_sec} is already assigned to coordinator: {existing_other.teacher.username}.")
                return render(request, 'add_teacher.html', {'dept_choices': dept_choices})

        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.is_staff = True
            user.save()
            TeacherProfile.objects.create(
                user=user,
                designation=designation,
                department=department
            )

            if is_coord:
                ClassCoordinator.objects.create(
                    teacher=user,
                    department=coord_dept,
                    year=coord_year,
                    section=coord_sec
                )

            messages.success(request, f"Teacher '{username}' added to {department} department!")
            return redirect('admin_dashboard')
        except Exception as e:
            messages.error(request, f"Error adding teacher: {e}")
            return render(request, 'add_teacher.html', {'dept_choices': dept_choices})

    return render(request, 'add_teacher.html', {'dept_choices': dept_choices})

@user_passes_test(lambda u: u.is_superuser)
def add_student(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        roll_number = request.POST.get('roll_number')
        department = request.POST.get('department')
        year = request.POST.get('year')
        section = request.POST.get('section')
        images = request.FILES.getlist('images')
        
        if Student.objects.filter(roll_number=roll_number).exists():
            messages.error(request, "Student with this Roll Number already exists.")
            return redirect('add_student')
            
        # Create User for student
        try:
            user = User.objects.create_user(username=roll_number, password=roll_number)
            student = Student.objects.create(
                name=name, 
                roll_number=roll_number, 
                user=user, 
                plain_password=roll_number,
                department=department,
                year=year,
                section=section
            )
        except Exception as e:
            messages.error(request, f"Error creating user: {e}")
            return redirect('add_student')
        
        # Process images if uploaded
        count = 0
        if images:
            folder_name = f"{roll_number}_{name}"
            safe_dept = "".join([c for c in department if c.isalnum() or c in (' ', '_', '-')]).strip()
            safe_year = "".join([c for c in str(year) if c.isalnum()]).strip()
            safe_section = "".join([c for c in str(section) if c.isalnum()]).strip()
            
            save_dir = os.path.join(settings.DATASET_DIR, safe_dept, safe_year, safe_section, folder_name)
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
                
            fs = FileSystemStorage()
            for img in images:
                filename = fs.save(img.name, img)
                temp_path = fs.path(filename)
                
                if detect_and_crop_face(temp_path, save_dir, folder_name):
                    count += 1
                
                fs.delete(filename)
                
            if count >= 5:
                student.is_registered = True
                student.save()
                train_model()
                messages.success(request, f"Student added with {count} face images. Model retrained.")
            else:
                messages.success(request, f"Student added with {count} face images. Student needs to complete registration to activate.")
        else:
            messages.success(request, "Student account created successfully. The student can now activate their account and register their face.")
            
        return redirect('admin_dashboard')
        
    return render(request, 'add_student.html')

@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def download_attendance(request):
    import openpyxl
    from openpyxl.styles import Font
    from django.http import HttpResponse

    # Get filter parameters
    branch = request.GET.get('branch')
    year = request.GET.get('year')
    section = request.GET.get('section')

    # Fetch Data (Reuse logic)
    if branch and year and section:
        students = Student.objects.filter(
            department__icontains=branch,
            year__icontains=year,
            section__icontains=section
        ).order_by('roll_number')
    else:
        students = Student.objects.none()

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Headers
    headers = ['Roll Number', 'Name', 'Department', 'Year', 'Section', 'Total Classes', 'Present', 'Attendance %']
    ws.append(headers)

    # Style Header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Determine if user is a teacher and filter subjects
    assigned_subjects = None
    if not request.user.is_superuser and request.user.is_staff:
        assigned_subjects = TeacherSubject.objects.filter(teacher=request.user).values_list('subject', flat=True)

    # Add Data
    for student in students:
        records = AttendanceRecord.objects.filter(student=student)
        if assigned_subjects is not None:
            records = records.filter(subject__in=assigned_subjects)
            
        total_classes = records.count()
        present_count = records.filter(status='Present').count()
        if total_classes > 0:
            percentage = round((present_count / total_classes) * 100, 1)
        else:
            percentage = 0
        
        ws.append([
            student.roll_number,
            student.name,
            student.department,
            student.year,
            student.section,
            total_classes,
            present_count,
            f"{percentage}%"
        ])

    # Create Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Attendance_Report_{branch}_{year}_{section}.xlsx"'

    wb.save(response)
    return response

@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def manage_students(request):
    # Get filter parameters
    branch = request.GET.get('branch')
    year = request.GET.get('year')
    section = request.GET.get('section')
    
    # Base QuerySet
    if branch and year and section:
         students = Student.objects.filter(
            department__icontains=branch,
            year__icontains=year,
            section__icontains=section
        ).order_by('roll_number')
    else:
        # Default behavior if no filter found
        if request.user.is_superuser:
            students = Student.objects.all().order_by('-created_at') # Admin sees all
        else:
            students = Student.objects.none() # Teacher sees none
            
    # Determine if user is a teacher and filter subjects
    assigned_subjects = None
    if not request.user.is_superuser and request.user.is_staff:
        assigned_subjects = TeacherSubject.objects.filter(teacher=request.user).values_list('subject', flat=True)

    # Calculate attendance for each student (common logic)
    for student in students:
        records = AttendanceRecord.objects.filter(student=student)
        if assigned_subjects is not None:
            records = records.filter(subject__in=assigned_subjects)
            
        total_classes = records.count()
        present_count = records.filter(status='Present').count()
        
        if total_classes > 0:
            percentage = round((present_count / total_classes) * 100, 1)
        else:
            percentage = 0
            
        student.attendance_percentage = percentage
        student.total_classes = total_classes
        student.present_count = present_count
        
    return render(request, 'manage_students.html', {'students': students})

@user_passes_test(lambda u: u.is_superuser)
def delete_student(request, student_id):
    if request.method == 'POST':
        student = get_object_or_404(Student, id=student_id)
        user = student.user
        name = student.name
        
        # Delete student folder if exists
        try:
            folder_name = f"{student.roll_number}_{student.name}"
            save_dir = os.path.join(settings.DATASET_DIR, folder_name)
            if os.path.exists(save_dir):
                import shutil
                shutil.rmtree(save_dir)
        except Exception as e:
            print(f"Error deleting folder: {e}")

        # Delete student (this cascades to attendance records)
        student.delete()
        
        # Delete associated user
        if user:
            user.delete()
            
        messages.success(request, f"Student {name} has been deleted successfully.")
        return redirect('manage_students')
    
    return redirect('manage_students')

@user_passes_test(lambda u: u.is_superuser)
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        roll_number = request.POST.get('roll_number')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        department = request.POST.get('department')
        year = request.POST.get('year')
        section = request.POST.get('section')
        password = request.POST.get('password')
        
        # Update fields
        student.name = name
        student.roll_number = roll_number
        student.email = email
        student.phone_number = phone_number
        student.department = department
        student.year = year
        student.section = section
        
        # Update password if provided
        if password:
            student.plain_password = password
            # Also update the User object if it exists
            if student.user:
                student.user.set_password(password)
                student.user.save()
        
        # Sync other details to User object if it exists
        if student.user:
            student.user.email = email
            student.user.first_name = name
            student.user.save()
        
        student.save()
        messages.success(request, f"Student {student.name} updated successfully.")
        return redirect('manage_students')
        
    return render(request, 'edit_student.html', {'student': student})

@user_passes_test(lambda u: u.is_superuser)
def manage_teachers(request):
    from .models import TeacherProfile
    dept_choices = TeacherProfile.DEPARTMENT_CHOICES

    # Filter by department if query param provided
    selected_dept = request.GET.get('department', '')

    teachers_qs = User.objects.filter(is_staff=True, is_superuser=False).order_by('username')

    # Ensure every teacher has a TeacherProfile
    for t in teachers_qs:
        TeacherProfile.objects.get_or_create(user=t)

    if selected_dept:
        teachers_qs = teachers_qs.filter(teacher_profile__department=selected_dept)

    # Group teachers by department for display
    from collections import OrderedDict
    dept_label_map = dict(dept_choices)
    grouped = OrderedDict()
    for t in teachers_qs:
        dept_key = t.teacher_profile.department
        dept_label = dept_label_map.get(dept_key, dept_key)
        grouped.setdefault(dept_label, []).append(t)

    return render(request, 'manage_teachers.html', {
        'teachers':      teachers_qs,
        'grouped':       grouped,
        'dept_choices':  dept_choices,
        'selected_dept': selected_dept,
    })

@user_passes_test(lambda u: u.is_superuser)
def edit_teacher(request, teacher_id):
    from .models import TeacherProfile
    teacher = get_object_or_404(User, id=teacher_id, is_staff=True, is_superuser=False)
    profile, _ = TeacherProfile.objects.get_or_create(user=teacher)
    dept_choices = TeacherProfile.DEPARTMENT_CHOICES
    
    coordinator = ClassCoordinator.objects.filter(teacher=teacher).first()

    if request.method == 'POST':
        username    = request.POST.get('username', '').strip()
        email       = request.POST.get('email', '').strip()
        password    = request.POST.get('password', '').strip()
        designation = request.POST.get('designation', 'asst_prof')
        department  = request.POST.get('department', 'CSE')

        # Coordinator post fields
        is_coord    = request.POST.get('is_coordinator') == 'on'
        coord_dept  = request.POST.get('coord_department', '').strip()
        coord_year  = request.POST.get('coord_year', '').strip()
        coord_sec   = request.POST.get('coord_section', '').strip()

        if User.objects.filter(username=username).exclude(pk=teacher.pk).exists():
            messages.error(request, "Username already taken by another user.")
            return redirect('edit_teacher', teacher_id=teacher_id)

        if is_coord:
            if not coord_dept or not coord_year or not coord_sec:
                messages.error(request, "All coordinator class fields (Department, Year, Section) are required when 'Assign as Class Coordinator' is checked.")
                return redirect('edit_teacher', teacher_id=teacher_id)

            existing_other = ClassCoordinator.objects.filter(
                department=coord_dept,
                year=coord_year,
                section=coord_sec
            ).exclude(teacher=teacher).first()
            if existing_other:
                messages.error(request, f"Class {coord_dept} Year {coord_year}-{coord_sec} is already assigned to coordinator: {existing_other.teacher.username}.")
                return redirect('edit_teacher', teacher_id=teacher_id)

            ClassCoordinator.objects.update_or_create(
                teacher=teacher,
                defaults={
                    'department': coord_dept,
                    'year': coord_year,
                    'section': coord_sec
                }
            )
        else:
            ClassCoordinator.objects.filter(teacher=teacher).delete()

        teacher.username = username
        teacher.email    = email
        if password:
            teacher.set_password(password)
        teacher.save()

        profile.designation = designation
        profile.department  = department
        profile.save()

        messages.success(request, f"Teacher '{username}' updated successfully!")
        return redirect('manage_teachers')

    return render(request, 'edit_teacher.html', {
        'teacher':      teacher,
        'profile':      profile,
        'dept_choices': dept_choices,
        'coordinator':  coordinator,
    })

@user_passes_test(lambda u: u.is_superuser)
def delete_teacher(request, teacher_id):
    if request.method == 'POST':
        teacher = get_object_or_404(User, id=teacher_id, is_staff=True, is_superuser=False)
        name = teacher.username
        teacher.delete()
        messages.success(request, f"Teacher '{name}' deleted successfully.")
    return redirect('manage_teachers')

@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    total_students = Student.objects.count()
    total_teachers = User.objects.filter(is_staff=True, is_superuser=False).count()
    total_attendance = AttendanceRecord.objects.count()
    
    # Get recent attendance
    recent_attendance = AttendanceRecord.objects.all().order_by('-date', '-time')[:5]
    
    # Assessment Requests
    pending_requests = AssessmentRequest.objects.filter(status='Pending').order_by('-created_at')
    pending_count = pending_requests.count()

    # Accessory Requests
    pending_accessory_requests = AccessoryRequest.objects.filter(status='Pending').order_by('-created_at')
    pending_accessory_count = pending_accessory_requests.count()

    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_attendance': total_attendance,
        'recent_attendance': recent_attendance,
        'pending_requests': pending_requests,
        'pending_requests_count': pending_count,
        'pending_accessory_requests': pending_accessory_requests,
        'pending_accessory_count': pending_accessory_count,
    }
    return render(request, 'admin_dashboard.html', context)


@user_passes_test(lambda u: u.is_superuser)
def review_assessment_request(request, request_id):
    """Admin: Approve or Reject a teacher assessment request."""
    req = get_object_or_404(AssessmentRequest, id=request_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        admin_remarks = request.POST.get('admin_remarks', '')
        if action == 'approve':
            req.status = 'Approved'
            req.admin_remarks = admin_remarks
            req.save()
            messages.success(request, f"Request '{req.title}' approved.")
        elif action == 'reject':
            req.status = 'Rejected'
            req.admin_remarks = admin_remarks
            req.save()
            messages.warning(request, f"Request '{req.title}' rejected.")
    return redirect('admin_dashboard')

@login_required
def profile(request):
    is_student = hasattr(request.user, 'student')
    student = request.user.student if is_student else None

    if request.method == 'POST':
        user = request.user
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        
        try:
            user.save()
            
            if is_student and student:
                student.name = f"{first_name} {last_name}".strip()
                student.email = email
                student.phone_number = request.POST.get('phone_number', '').strip()
                student.address = request.POST.get('address', '').strip()
                
                dob = request.POST.get('date_of_birth')
                if dob:
                    student.date_of_birth = dob
                
                if 'profile_pic' in request.FILES:
                    student.profile_pic = request.FILES['profile_pic']
                
                student.save()
                
            messages.success(request, "Profile updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating profile: {e}")
        return redirect('profile')

    user = request.user
    if is_student and student:
        if not user.first_name and not user.last_name:
            name_parts = student.name.split(' ', 1)
            user.first_name = name_parts[0]
            user.last_name = name_parts[1] if len(name_parts) > 1 else ''
        if not user.email:
            user.email = student.email or ''

    return render(request, 'profile.html', {
        'user': user,
        'is_student': is_student,
        'student': student
    })

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            
            # Update plain password for display in admin dashboard
            if hasattr(user, 'student'):
                try:
                    student = user.student
                    student.plain_password = form.cleaned_data['new_password1']
                    student.save()
                    print(f"DEBUG: Updated plain_password for {student.roll_number} to {student.plain_password}")
                except Exception as e:
                    print(f"DEBUG: Error updating plain_password: {e}")
            else:
                 print(f"DEBUG: User {user.username} has no student profile.")
                
            messages.success(request, 'Your password was successfully updated!')
            # Redirect to appropriate dashboard based on role
            if user.is_superuser:
                return redirect('admin_dashboard')
            elif user.is_staff:
                return redirect('teacher_dashboard')
            else:
                return redirect('student_dashboard')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'change_password.html', {'form': form})

@login_required
def student_dashboard(request):
    if not hasattr(request.user, 'student'):
         # Fallback if user is not linked to a student profile
         return render(request, 'student_portal/dashboard.html', {'error': 'No student profile found.'})
    
    student = request.user.student
    
    # Handle Profile Update
    if request.method == 'POST':
        try:
            student.email = request.POST.get('email')
            student.phone_number = request.POST.get('phone_number')
            student.department = request.POST.get('department')
            student.address = request.POST.get('address')
            
            dob = request.POST.get('date_of_birth')
            if dob:
                student.date_of_birth = dob
            
            if 'profile_pic' in request.FILES:
                student.profile_pic = request.FILES['profile_pic']
                
            student.save()
            messages.success(request, "Profile updated successfully!")
            return redirect('student_dashboard')
        except Exception as e:
            messages.error(request, f"Error updating profile: {e}")
    
    today = datetime.date.today()
    start_week = today - datetime.timedelta(days=today.weekday())
    end_week = start_week + datetime.timedelta(days=6)
    
    # Calculate attendance stats
    current_attendance_count = AttendanceRecord.objects.filter(student=student, status='Present').count()
    
    # Subject-wise Attendance Logic
    all_records = AttendanceRecord.objects.filter(student=student)
    subject_stats = {}
    
    for record in all_records:
        # Use subject from record, default to "General" if None
        subj = record.subject if record.subject else "General"
        
        if subj not in subject_stats:
            subject_stats[subj] = {'present': 0, 'total': 0}
            
        subject_stats[subj]['total'] += 1
        if record.status == 'Present':
            subject_stats[subj]['present'] += 1
            
    subject_attendance = []
    for subj, stats in subject_stats.items():
        total = stats['total']
        present = stats['present']
        percentage = (present / total * 100) if total > 0 else 0
        subject_attendance.append({
            'subject': subj,
            'present': present,
            'total': total,
            'percentage': round(percentage, 1)
        })
    
    # Overall Statistics
    total_classes = all_records.count()
    overall_percentage = (current_attendance_count / total_classes * 100) if total_classes > 0 else 0
    
    # Prediction Logic
    # Estimate weekly classes from Timetable count or default
    weekly_classes_count = TimeTable.objects.count()
    if weekly_classes_count == 0:
        weekly_classes_count = 15 # Default assumption if no timetable
        
    # Predict Next Week (Assuming full attendance)
    pred_week_present = current_attendance_count + weekly_classes_count
    pred_week_total = total_classes + weekly_classes_count
    pred_week_pct = (pred_week_present / pred_week_total * 100) if pred_week_total > 0 else 0
    
    # Predict Next Month (Assuming full attendance ~4 weeks)
    pred_month_present = current_attendance_count + (weekly_classes_count * 4)
    pred_month_total = total_classes + (weekly_classes_count * 4)
    pred_month_pct = (pred_month_present / pred_month_total * 100) if pred_month_total > 0 else 0

    records = AttendanceRecord.objects.filter(student=student).order_by('-date', '-time')[:10]
    
    context = {
        'student': student,
        'current_attendance_count': current_attendance_count,
        'total_classes': total_classes,
        'overall_percentage': round(overall_percentage, 1),
        'records': records,
        'subject_attendance': subject_attendance,
        'prediction_week': round(pred_week_pct, 1),
        'prediction_month': round(pred_month_pct, 1),
    }
    return render(request, 'student_portal/dashboard.html', context)

@login_required
def teacher_dashboard(request):
    # Fetch subjects assigned to this teacher
    today_weekday = datetime.datetime.now().weekday()
    assigned_classes = TeacherSubject.objects.filter(teacher=request.user, day=today_weekday)
    
    # Filter by Year
    year_query = request.GET.get('year')
    if year_query:
        assigned_classes = assigned_classes.filter(year__icontains=year_query)
        
    # Filter by Section
    section_query = request.GET.get('section')
    if section_query:
        assigned_classes = assigned_classes.filter(section__icontains=section_query)
        
    # Calculate Stats for Dashboard
    # Total assigned classes
    total_classes_count = assigned_classes.count()
    
    # Total students in relevant years/sections (Approximate)
    # Get set of (year, section) tuples
    class_specs = assigned_classes.values_list('year', 'section').distinct()
    
    total_students_count = 0
    relevant_students_ids = set()
    
    for year, section in class_specs:
        q = Student.objects.filter(year__icontains=year)
        if section:
            q = q.filter(section__icontains=section)
            # If section name is a department code, ensure student's department also matches to avoid cross-branch matching
            dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
            if section.upper() in dept_codes:
                q = q.filter(department__icontains=section)
        ids = q.values_list('id', flat=True)
        relevant_students_ids.update(ids)
        
    total_students_count = len(relevant_students_ids)
    
    # Recent Attendance for these classes/students
    recent_attendance = AttendanceRecord.objects.filter(student__id__in=relevant_students_ids).order_by('-date', '-time')[:5]

    # Annotate classes with active status (is_active) based on current time
    current_time = datetime.datetime.now().time()
    display_classes = []
    for cls in assigned_classes:
        cls.is_active = False
        if cls.start_time and cls.end_time:
            # Check if current time is within class hours
            if cls.start_time <= cls.end_time:
                # Normal case: start <= now <= end
                if cls.start_time <= current_time <= cls.end_time:
                    cls.is_active = True
            else:
                # Midnight crossover: start > end (e.g. 23:00 to 01:00)
                # Active if now >= start (23:30) OR now <= end (00:30)
                if current_time >= cls.start_time or current_time <= cls.end_time:
                    cls.is_active = True
        display_classes.append(cls)

    return render(request, 'teacher_dashboard.html', {
        'assigned_classes': display_classes,
        'selected_year': year_query, 
        'selected_section': section_query,
        'total_classes_count': total_classes_count,
        'total_students_count': total_students_count,
        'recent_attendance': recent_attendance,
        'user': request.user
    })

@user_passes_test(lambda u: u.is_superuser)
def manage_teacher_subjects(request):
    subjects = TeacherSubject.objects.all().order_by('teacher__username', 'day')
    return render(request, 'manage_teacher_subjects.html', {'subjects': subjects})

@user_passes_test(lambda u: u.is_superuser)
def assign_teacher_subject(request):
    teachers = User.objects.filter(is_staff=True, is_superuser=False)
    
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher')
        subject = request.POST.get('subject')
        year = request.POST.get('year')
        section = request.POST.get('section')
        day = request.POST.get('day')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        
        # Convert empty strings to None
        if not start_time: start_time = None
        if not end_time: end_time = None
        
        try:
            teacher = User.objects.get(id=teacher_id)
            TeacherSubject.objects.create(
                teacher=teacher,
                subject=subject,
                year=year,
                section=section,
                day=day,
                start_time=start_time,
                end_time=end_time
            )
            messages.success(request, f"Assigned {subject} to {teacher.username} successfully.")
            return redirect('manage_teacher_subjects')
        except Exception as e:
            messages.error(request, f"Error assigning subject: {e}")
            
    days = TeacherSubject.DAYS_OF_WEEK
    return render(request, 'assign_teacher_subject.html', {'teachers': teachers, 'days': days})

@user_passes_test(lambda u: u.is_superuser)
def edit_teacher_subject(request, subject_id):
    subject_obj = get_object_or_404(TeacherSubject, id=subject_id)
    teachers = User.objects.filter(is_staff=True, is_superuser=False)
    
    if request.method == 'POST':
        try:
            teacher_id = request.POST.get('teacher')
            subject = request.POST.get('subject')
            year = request.POST.get('year')
            section = request.POST.get('section')
            day = request.POST.get('day')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            
            # Convert empty strings to None
            if not start_time: start_time = None
            if not end_time: end_time = None
            
            teacher = User.objects.get(id=teacher_id)
            
            # Update fields
            subject_obj.teacher = teacher
            subject_obj.subject = subject
            subject_obj.year = year
            subject_obj.section = section
            subject_obj.day = day
            subject_obj.start_time = start_time
            subject_obj.end_time = end_time
            subject_obj.save()
            
            messages.success(request, f"Updated assignment for {subject} successfully.")
            return redirect('manage_teacher_subjects')
        except Exception as e:
            messages.error(request, f"Error updating assignment: {e}")
            
    days = TeacherSubject.DAYS_OF_WEEK
    return render(request, 'edit_teacher_subject.html', {
        'submission': subject_obj,
        'teachers': teachers, 
        'days': days
    })

@user_passes_test(lambda u: u.is_superuser)
def delete_teacher_subject(request, subject_id):
    if request.method == 'POST':
        try:
            subject = TeacherSubject.objects.get(id=subject_id)
            name = subject.subject
            teacher = subject.teacher.username
            subject.delete()
            messages.success(request, f"Removed assignment: {name} from {teacher}")
        except TeacherSubject.DoesNotExist:
            messages.error(request, "Subject assignment not found.")
        except Exception as e:
            messages.error(request, f"Error deleting assignment: {e}")
            
    return redirect('manage_teacher_subjects')

# --- Teacher Store / Assessment Request + Accessories ---

@login_required
def teacher_store(request):
    """Teacher Store: Unified page for assessment and accessory requests."""
    if not request.user.is_staff or request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('home')

    # Teacher's assigned subjects (for auto-fill dropdown)
    teacher_subjects = TeacherSubject.objects.filter(teacher=request.user).values_list('subject', flat=True).distinct()

    if request.method == 'POST':
        form_type = request.POST.get('form_type', 'assessment')

        if form_type == 'accessory':
            # ── Accessory Request ──
            accessory_type = request.POST.get('accessory_type', 'pen')
            quantity = request.POST.get('quantity', '1').strip() or '1'
            priority = request.POST.get('priority', 'medium')
            notes = request.POST.get('notes', '').strip()
            try:
                quantity = max(1, int(quantity))
            except ValueError:
                quantity = 1
            AccessoryRequest.objects.create(
                teacher=request.user,
                accessory_type=accessory_type,
                quantity=quantity,
                priority=priority,
                notes=notes,
            )
            acc_label = dict(AccessoryRequest.ACCESSORY_TYPES).get(accessory_type, accessory_type)
            messages.success(request, f"Accessory request for '{acc_label}' (x{quantity}) submitted!")
            return redirect('teacher_store')

        else:
            # ── Assessment Request ──
            subject = request.POST.get('subject', '').strip()
            assessment_type = request.POST.get('assessment_type', 'question_paper')
            title = request.POST.get('title', '').strip()
            description = request.POST.get('description', '').strip()
            year = request.POST.get('year', '').strip()
            section = request.POST.get('section', '').strip()

            if not subject or not title:
                messages.error(request, 'Subject and Title are required fields.')
            else:
                AssessmentRequest.objects.create(
                    teacher=request.user,
                    subject=subject,
                    assessment_type=assessment_type,
                    title=title,
                    description=description,
                    year=year,
                    section=section,
                )
                messages.success(request, f"Assessment request for '{title}' submitted successfully!")
                return redirect('teacher_store')

    my_requests = AssessmentRequest.objects.filter(teacher=request.user)
    my_accessory_requests = AccessoryRequest.objects.filter(teacher=request.user)
    assessment_types = AssessmentRequest.ASSESSMENT_TYPES
    accessory_types = AccessoryRequest.ACCESSORY_TYPES
    priority_choices = AccessoryRequest.PRIORITY_CHOICES

    return render(request, 'teacher_portal/teacher_store.html', {
        'my_requests': my_requests,
        'my_accessory_requests': my_accessory_requests,
        'teacher_subjects': teacher_subjects,
        'assessment_types': assessment_types,
        'accessory_types': accessory_types,
        'priority_choices': priority_choices,
    })


@login_required
def cancel_assessment_request(request, request_id):
    """Teacher: Cancel a pending assessment request."""
    req = get_object_or_404(AssessmentRequest, id=request_id, teacher=request.user)
    if req.status == 'Pending':
        req.delete()
        messages.success(request, f"Request '{req.title}' has been cancelled.")
    else:
        messages.error(request, "Only pending requests can be cancelled.")
    return redirect('teacher_store')


@login_required
def cancel_accessory_request(request, request_id):
    """Teacher: Cancel a pending accessory request."""
    req = get_object_or_404(AccessoryRequest, id=request_id, teacher=request.user)
    if req.status == 'Pending':
        req.delete()
        messages.success(request, f"Accessory request cancelled.")
    else:
        messages.error(request, "Only pending requests can be cancelled.")
    return redirect('teacher_store')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def review_accessory_request(request, request_id):
    """Admin: Approve or reject an accessory request."""
    req = get_object_or_404(AccessoryRequest, id=request_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        admin_remarks = request.POST.get('admin_remarks', '').strip()
        if action == 'approve':
            req.status = 'Approved'
        elif action == 'reject':
            req.status = 'Rejected'
        if admin_remarks:
            req.admin_remarks = admin_remarks
        req.save()
        messages.success(request, f"Accessory request has been {req.status.lower()}.")
    return redirect('admin_dashboard')


# --- Notification Logic ---

def create_notification(student, message, notif_type='Attendance'):
    """Helper to create a notification for a student."""
    try:
        Notification.objects.create(
            recipient=student,
            message=message,
            notification_type=notif_type
        )
    except Exception as e:
        print(f"Error creating notification: {e}")


def send_store_notification(recipients, message, store_request=None, notif_type='general'):
    """Send a StoreNotification to one or more User recipients."""
    if isinstance(recipients, User):
        recipients = [recipients]
    for user in recipients:
        if user:
            try:
                StoreNotification.objects.create(
                    recipient=user,
                    store_request=store_request,
                    message=message,
                    notif_type=notif_type,
                )
            except Exception as e:
                print(f"[StoreNotification] Error for {user}: {e}")


def _get_dept_hod_user(department):
    """Return HOD User for a given department, or None."""
    return User.objects.filter(
        teacher_profile__department=department,
        teacher_profile__designation='hod',
        is_staff=True,
        is_superuser=False,
    ).first()


def _get_store_head_users():
    """Return all Store Head Users."""
    return list(User.objects.filter(store_profile__role='head'))

@login_required
def get_notifications(request):
    """API to fetch unread notifications for the logged-in student."""
    if hasattr(request.user, 'student'):
        student = request.user.student
        # Get unread notifications
        notifications = Notification.objects.filter(recipient=student, is_read=False)[:10]
        data = [{
            'id': n.id,
            'message': n.message,
            'created_at': n.created_at.strftime('%H:%M %d-%m'),
            'type': n.notification_type
        } for n in notifications]
        return JsonResponse({'status': 'success', 'notifications': data, 'count': len(data)})
    return JsonResponse({'status': 'error', 'message': 'Not a student'}, status=403)

@login_required
@csrf_exempt
def mark_notification_read(request, notif_id):
    """API to mark a notification as read."""
    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notif_id, recipient__user=request.user)
            notification.is_read = True
            notification.save()
            return JsonResponse({'status': 'success'})
        except Notification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Notification not found'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


# ═══════════════════════════════════════════════
#  STORE MANAGEMENT VIEWS
# ═══════════════════════════════════════════════

@user_passes_test(lambda u: u.is_superuser)
def admin_store_dashboard(request):
    """Admin: Full store management — staff + all requests with item details."""
    store_staff   = StoreStaff.objects.select_related('user').order_by('role', 'user__username')
    status_filter = request.GET.get('status', '')

    requests_qs = StoreRequest.objects.select_related(
        'requested_by', 'fulfilled_by__user'
    ).prefetch_related('items').all()

    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)

    # Only show users who are NOT teachers, NOT students, NOT superusers, NOT already store staff
    from django.db.models import Q
    all_users = User.objects.filter(is_active=True).exclude(
        id__in=StoreStaff.objects.values_list('user_id', flat=True)
    ).exclude(
        Q(is_staff=True) | Q(is_superuser=True) | Q(student__isnull=False)
    ).order_by('username')

    context = {
        'store_staff':    store_staff,
        'store_requests': requests_qs,
        'status_filter':  status_filter,
        'STATUS_CHOICES': StoreRequest.STATUS_CHOICES,
        'all_users':      all_users,
    }
    return render(request, 'store_dashboard.html', context)


@user_passes_test(lambda u: u.is_superuser)
def add_store_staff(request):
    """Admin: Add a user as Store Head or Store Staff."""
    if request.method == 'POST':
        action     = request.POST.get('action')
        role       = request.POST.get('role', 'staff')
        phone      = request.POST.get('phone', '').strip()
        department = request.POST.get('department', '').strip()

        if action == 'add_existing':
            user_id = request.POST.get('user_id')
            try:
                user = User.objects.get(id=user_id)
                StoreStaff.objects.create(user=user, role=role, phone=phone, department=department)
                messages.success(request, f"'{user.username}' added as {role}!")
            except Exception as e:
                messages.error(request, f"Error: {e}")

        elif action == 'create_new':
            username = request.POST.get('username', '').strip()
            email    = request.POST.get('email', '').strip()
            password = request.POST.get('password', '').strip()
            if User.objects.filter(username=username).exists():
                messages.error(request, "Username already exists.")
            else:
                try:
                    user = User.objects.create_user(username=username, email=email, password=password)
                    StoreStaff.objects.create(user=user, role=role, phone=phone, department=department)
                    messages.success(request, f"Store staff '{username}' created!")
                except Exception as e:
                    messages.error(request, f"Error: {e}")

    return redirect('admin_store_dashboard')


@user_passes_test(lambda u: u.is_superuser)
def delete_store_staff(request, staff_id):
    """Admin: Remove a user from store staff."""
    if request.method == 'POST':
        staff = get_object_or_404(StoreStaff, id=staff_id)
        name  = staff.user.username
        staff.delete()
        messages.success(request, f"'{name}' removed from store staff.")
    return redirect('admin_store_dashboard')


@user_passes_test(lambda u: u.is_superuser)
def edit_store_staff(request, staff_id):
    """Admin: Edit store staff details."""
    staff = get_object_or_404(StoreStaff, id=staff_id)
    if request.method == 'POST':
        role       = request.POST.get('role', staff.role)
        phone      = request.POST.get('phone', '').strip()
        department = request.POST.get('department', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '').strip()

        staff.role       = role
        staff.phone      = phone
        staff.department = department
        staff.save()

        if email:
            staff.user.email = email
        if password:
            staff.user.set_password(password)
        staff.user.save()

        messages.success(request, f"'{staff.user.username}' updated successfully!")
        return redirect('admin_store_dashboard')

    return render(request, 'edit_store_staff.html', {'staff': staff})


@user_passes_test(lambda u: u.is_superuser)
def update_store_request(request, request_id):
    """Admin: Update item fulfilment quantities + overall status."""
    store_req = get_object_or_404(StoreRequest, id=request_id)
    if request.method == 'POST':
        old_status = store_req.status

        # Update each item's quantity_provided
        for item in store_req.items.all():
            key     = f'qty_provided_{item.id}'
            new_qty = request.POST.get(key)
            if new_qty is not None:
                try:
                    item.quantity_provided = max(0, int(new_qty))
                    item.save()
                except ValueError:
                    pass

        # Manual status override (optional)
        manual_status   = request.POST.get('status', '').strip()
        remarks         = request.POST.get('remarks', '').strip()
        fulfilled_by_id = request.POST.get('fulfilled_by', '').strip()

        # Auto-compute status from items
        items = list(store_req.items.all())
        if items:
            if all(i.is_fulfilled for i in items):
                store_req.status = 'fulfilled'
            elif any(i.quantity_provided > 0 for i in items):
                store_req.status = 'partial'
            else:
                store_req.status = store_req.status  # keep current
        elif manual_status:
            store_req.status = manual_status

        store_req.remarks = remarks
        if fulfilled_by_id:
            try:
                store_req.fulfilled_by = StoreStaff.objects.get(id=fulfilled_by_id)
            except StoreStaff.DoesNotExist:
                pass
        store_req.save()

        # ── Send notifications to all stakeholders ──────────────────────
        new_status      = store_req.status
        status_label    = dict(StoreRequest.STATUS_CHOICES).get(new_status, new_status)
        req_title       = store_req.title
        teacher_user    = store_req.requested_by
        notif_message   = (
            f"[Admin Update] Request '{req_title}' status → '{status_label}'. "
            f"Remarks: {remarks}" if remarks else
            f"[Admin Update] Request '{req_title}' status → '{status_label}'."
        )

        notify_users = [teacher_user]   # always notify the requesting teacher

        # Notify HOD if request came from a teacher with a dept
        try:
            dept = teacher_user.teacher_profile.department
            hod_user = _get_dept_hod_user(dept)
            if hod_user and hod_user != teacher_user:
                notify_users.append(hod_user)
        except Exception:
            pass

        # Notify Store Heads
        notify_users.extend(_get_store_head_users())

        # Notify assigned store staff (if any)
        if store_req.assigned_to:
            notify_users.append(store_req.assigned_to.user)
        if store_req.fulfilled_by and store_req.fulfilled_by != store_req.assigned_to:
            notify_users.append(store_req.fulfilled_by.user)

        # Deduplicate
        seen = set()
        unique_users = []
        for u in notify_users:
            if u and u.id not in seen:
                seen.add(u.id)
                unique_users.append(u)

        send_store_notification(
            unique_users, notif_message,
            store_request=store_req, notif_type='status_update'
        )

        messages.success(request, "Request updated and all stakeholders notified!")
    return redirect('admin_store_dashboard')


@login_required
def submit_store_request(request):
    """Teacher / any staff: Submit a new store request with multiple items."""
    if request.method == 'POST':
        title      = request.POST.get('title', '').strip()
        notes      = request.POST.get('notes', '').strip()
        item_names = request.POST.getlist('item_name')
        item_qtys  = request.POST.getlist('item_qty')

        if not title or not any(n.strip() for n in item_names):
            messages.error(request, "Please provide a title and at least one item.")
            return render(request, 'submit_store_request.html')

        store_req = StoreRequest.objects.create(
            requested_by=request.user,
            title=title,
            notes=notes,
        )
        for name, qty in zip(item_names, item_qtys):
            name = name.strip()
            if name:
                try:
                    qty_int = max(1, int(qty))
                except (ValueError, TypeError):
                    qty_int = 1
                StoreRequestItem.objects.create(
                    request=store_req,
                    item_name=name,
                    quantity_requested=qty_int
                )
        messages.success(request, f"Store request '{title}' submitted!")
        return redirect('teacher_dashboard')

    return render(request, 'submit_store_request.html')


# ───────────────────────────────────────────────────────
#  STORE WORKFLOW — HOD → Store Head → Store Staff
# ───────────────────────────────────────────────────────

def _get_hod_profile(user):
    """Return TeacherProfile if user is HOD, else None."""
    try:
        p = user.teacher_profile
        return p if p.designation == 'hod' else None
    except Exception:
        return None

def _get_store_staff_profile(user):
    """Return StoreStaff profile if user is store staff/head, else None."""
    try:
        return user.store_profile
    except Exception:
        return None


# ── HOD: see & approve requests from their dept ──────────────────────
@login_required
def hod_store_requests(request):
    """HOD sees pending store requests from teachers in their department."""
    hod_profile = _get_hod_profile(request.user)
    if not hod_profile:
        messages.error(request, "Access denied — HOD only.")
        return redirect('teacher_dashboard')

    dept = hod_profile.department
    # Teachers in same dept (not HOD themselves)
    dept_teacher_ids = User.objects.filter(
        teacher_profile__department=dept,
        is_staff=True, is_superuser=False
    ).values_list('id', flat=True)

    pending   = StoreRequest.objects.filter(
        requested_by__in=dept_teacher_ids, status='pending_hod'
    ).prefetch_related('items').select_related('requested_by')

    reviewed  = StoreRequest.objects.filter(
        requested_by__in=dept_teacher_ids
    ).exclude(status='pending_hod').prefetch_related('items').select_related(
        'requested_by', 'hod_approved_by', 'assigned_to__user', 'fulfilled_by__user'
    )

    return render(request, 'hod_store_requests.html', {
        'pending':  pending,
        'reviewed': reviewed,
        'dept':     dept,
    })


@login_required
def hod_review_store_request(request, request_id):
    """HOD approves or rejects a store request."""
    hod_profile = _get_hod_profile(request.user)
    if not hod_profile:
        return redirect('teacher_dashboard')

    store_req   = get_object_or_404(StoreRequest, id=request_id, status='pending_hod')
    if request.method == 'POST':
        action      = request.POST.get('action')
        hod_remarks = request.POST.get('hod_remarks', '').strip()

        store_req.hod_approved_by  = request.user
        store_req.hod_approved_at  = timezone.now()
        store_req.hod_remarks      = hod_remarks

        if action == 'approve':
            store_req.status = 'pending_store'
            action_text = 'approved'
            messages.success(request, f"Request '{store_req.title}' approved → sent to Store Head.")
        else:
            store_req.status = 'hod_rejected'
            action_text = 'rejected'
            messages.warning(request, f"Request '{store_req.title}' rejected.")

        store_req.save()

        # Notify teacher and store heads
        hod_name = request.user.get_full_name() or request.user.username
        notif_msg = (
            f"Your request '{store_req.title}' has been {action_text} by HOD ({hod_name})."
            + (f" Remarks: {hod_remarks}" if hod_remarks else "")
        )
        notify_list = [store_req.requested_by] + _get_store_head_users()
        send_store_notification(notify_list, notif_msg, store_request=store_req, notif_type='hod_action')

    return redirect('hod_store_requests')

# ───────────────────────────────────────────────────────
#  COURSE MATERIALS (ASSIGNMENTS & NOTES)
# ───────────────────────────────────────────────────────

@login_required
def teacher_materials(request):
    """Teacher: Upload and manage Assignments and Notes."""
    if not request.user.is_staff or request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('home')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        material_type = request.POST.get('material_type', 'assignment')
        subject = request.POST.get('subject', '').strip()
        year = request.POST.get('year', '').strip()
        section = request.POST.get('section', '').strip()
        file = request.FILES.get('file')

        due_date_str = request.POST.get('due_date', '').strip()
        due_date = None
        if due_date_str:
            try:
                due_date = timezone.datetime.fromisoformat(due_date_str)
                if timezone.is_naive(due_date):
                    due_date = timezone.make_aware(due_date)
            except ValueError:
                pass

        if not title or not file:
            messages.error(request, 'Title and File are required.')
        else:
            try:
                CourseMaterial.objects.create(
                    teacher=request.user,
                    title=title,
                    description=description,
                    material_type=material_type,
                    subject=subject,
                    year=year,
                    section=section,
                    file=file,
                    due_date=due_date
                )
                messages.success(request, f"{material_type.capitalize()} '{title}' uploaded successfully!")
                return redirect('teacher_materials')
            except Exception as e:
                messages.error(request, f"Error uploading file: {e}")

    materials = CourseMaterial.objects.filter(teacher=request.user).annotate(submission_count=Count('submissions'))
    material_types = CourseMaterial.MATERIAL_TYPES
    teacher_subjects = TeacherSubject.objects.filter(teacher=request.user).values_list('subject', flat=True).distinct()

    return render(request, 'teacher_portal/manage_materials.html', {
        'materials': materials,
        'material_types': material_types,
        'teacher_subjects': teacher_subjects,
    })

@login_required
def delete_material(request, material_id):
    """Teacher: Delete an uploaded material."""
    if not request.user.is_staff:
        return redirect('home')
    
    material = get_object_or_404(CourseMaterial, id=material_id, teacher=request.user)
    if request.method == 'POST':
        title = material.title
        material.delete()
        messages.success(request, f"Deleted '{title}'.")
        
    return redirect('teacher_materials')


@login_required
def student_materials(request):
    """Student: View and download Assignments and Notes."""
    if not hasattr(request.user, 'student'):
        messages.error(request, 'Access denied. Student profile required.')
        return redirect('home')

    student = request.user.student
    
    materials_query = CourseMaterial.objects.all()
    
    # Filter by Year
    if student.year:
        materials_query = materials_query.filter(
            Q(year='') | Q(year__iexact=student.year)
        )
    
    # Filter by Section
    if student.section:
        materials_query = materials_query.filter(
            Q(section='') | Q(section__iexact=student.section)
        )

    # Allow custom filtering
    material_type = request.GET.get('type', '')
    if material_type:
        materials_query = materials_query.filter(material_type=material_type)

    materials = list(materials_query.select_related('teacher'))

    # Fetch submissions and late requests for student
    submissions = {s.assignment_id: s for s in StudentSubmission.objects.filter(student=student)}
    late_requests = {r.assignment_id: r for r in LateSubmissionRequest.objects.filter(student=student)}
    for m in materials:
        if m.material_type == 'assignment':
            m.submission = submissions.get(m.id)
            m.late_request = late_requests.get(m.id)

    return render(request, 'student_portal/student_materials.html', {
        'materials': materials,
        'selected_type': material_type,
    })


# ── Store Head: assign requests to staff ─────────────────────────────
@login_required
def store_head_dashboard(request):
    """Store Head sees pending_store requests and can assign them."""
    sp = _get_store_staff_profile(request.user)
    if not sp or sp.role != 'head':
        messages.error(request, "Access denied — Store Head only.")
        return redirect('login')

    pending_store = StoreRequest.objects.filter(
        status='pending_store'
    ).prefetch_related('items').select_related('requested_by', 'hod_approved_by')

    assigned = StoreRequest.objects.filter(
        status='assigned'
    ).prefetch_related('items').select_related('requested_by', 'assigned_to__user')

    done = StoreRequest.objects.filter(
        status__in=['delivered', 'partial', 'fulfilled', 'rejected']
    ).prefetch_related('items').select_related('requested_by', 'assigned_to__user')[:20]

    all_staff = StoreStaff.objects.filter(role='staff').select_related('user')

    return render(request, 'store_head_dashboard.html', {
        'pending_store': pending_store,
        'assigned':      assigned,
        'done':          done,
        'all_staff':     all_staff,
    })


@login_required
def assign_store_request(request, request_id):
    """Store Head assigns a request to a specific store staff member."""
    sp = _get_store_staff_profile(request.user)
    if not sp or sp.role != 'head':
        return redirect('login')

    store_req = get_object_or_404(StoreRequest, id=request_id, status='pending_store')
    if request.method == 'POST':
        staff_id = request.POST.get('staff_id', '').strip()
        remarks  = request.POST.get('remarks', '').strip()
        try:
            staff = StoreStaff.objects.get(id=staff_id)
            store_req.assigned_to  = staff
            store_req.assigned_at  = timezone.now()
            store_req.status       = 'assigned'
            store_req.remarks      = remarks
            store_req.save()
            messages.success(request, f"Assigned to {staff.user.username}!")

            # Notify store staff that they have a new task
            head_name = request.user.get_full_name() or request.user.username
            send_store_notification(
                staff.user,
                f"You have been assigned a new store request: '{store_req.title}' by Store Head ({head_name})."
                + (f" Note: {remarks}" if remarks else ""),
                store_request=store_req, notif_type='store_action'
            )
            # Notify the requesting teacher
            send_store_notification(
                store_req.requested_by,
                f"Your request '{store_req.title}' has been assigned to store staff ({staff.user.get_full_name() or staff.user.username}) and is being processed.",
                store_request=store_req, notif_type='store_action'
            )
        except (StoreStaff.DoesNotExist, ValueError):
            messages.error(request, "Staff not found or invalid ID.")
    return redirect('store_head_dashboard')


# ── Store Staff: see tasks & respond ─────────────────────────────────
@login_required
def store_staff_tasks(request):
    """Store Staff sees their assigned tasks."""
    sp = _get_store_staff_profile(request.user)
    if not sp:
        messages.error(request, "Access denied.")
        return redirect('login')

    active    = StoreRequest.objects.filter(
        assigned_to=sp, status='assigned'
    ).prefetch_related('items').select_related('requested_by')

    completed = StoreRequest.objects.filter(
        assigned_to=sp, status__in=['delivered', 'partial', 'fulfilled']
    ).prefetch_related('items').select_related('requested_by')

    return render(request, 'store_staff_tasks.html', {
        'active':    active,
        'completed': completed,
        'sp':        sp,
    })


@login_required
def store_staff_respond(request, request_id):
    """Store Staff submits fulfilment response with qty + expected delivery."""
    sp = _get_store_staff_profile(request.user)
    if not sp:
        return redirect('login')

    store_req = get_object_or_404(StoreRequest, id=request_id, assigned_to=sp)
    if request.method == 'POST':
        from datetime import date as date_cls
        staff_response    = request.POST.get('staff_response', '').strip()
        expected_delivery = request.POST.get('expected_delivery', '').strip()

        # Update each item's qty_provided
        for item in store_req.items.all():
            key     = f'qty_provided_{item.id}'
            new_qty = request.POST.get(key)
            if new_qty is not None:
                try:
                    item.quantity_provided = max(0, min(int(new_qty), item.quantity_requested))
                    item.save()
                except ValueError:
                    pass

        if expected_delivery:
            try:
                store_req.expected_delivery = date_cls.fromisoformat(expected_delivery)
            except ValueError:
                pass

        store_req.staff_response = staff_response
        store_req.fulfilled_by   = sp

        # Auto-compute status
        items = list(store_req.items.all())
        if items:
            if any(i.quantity_provided > 0 for i in items):
                store_req.status = 'delivered'
            else:
                store_req.status = 'rejected'

        store_req.save()
        messages.success(request, "Response submitted successfully!")

        # ── Notify teacher + store heads ─────────────────────────────────
        staff_name = request.user.get_full_name() or request.user.username
        status_label = dict(StoreRequest.STATUS_CHOICES).get(store_req.status, store_req.status)
        notif_msg = (
            f"Store Staff ({staff_name}) has fulfilled your request '{store_req.title}'. "
            f"Status: {status_label}. Please confirm receipt."
            + (f" Note: {staff_response}" if staff_response else "")
        )
        notify_list = [store_req.requested_by] + _get_store_head_users()
        # Also notify HOD
        try:
            dept = store_req.requested_by.teacher_profile.department
            hod_user = _get_dept_hod_user(dept)
            if hod_user:
                notify_list.append(hod_user)
        except Exception:
            pass
        send_store_notification(notify_list, notif_msg, store_request=store_req, notif_type='fulfillment')

    return redirect('store_staff_tasks')


@login_required
def teacher_confirm_receipt(request, request_id):
    """Teacher confirms receipt of store items."""
    store_req = get_object_or_404(StoreRequest, id=request_id, requested_by=request.user, status='delivered')
    
    if request.method == 'POST':
        teacher_remarks = request.POST.get('teacher_remarks', '').strip()
        
        # Update each item's qty_received
        for item in store_req.items.all():
            key = f'qty_received_{item.id}'
            new_qty = request.POST.get(key)
            if new_qty is not None:
                try:
                    item.quantity_received = max(0, min(int(new_qty), item.quantity_requested))
                    item.save()
                except ValueError:
                    pass
        
        store_req.teacher_remarks = teacher_remarks
        store_req.teacher_confirmed_at = timezone.now()
        
        # Auto-compute status
        items = list(store_req.items.all())
        if items:
            if all(i.quantity_received >= i.quantity_requested for i in items):
                store_req.status = 'fulfilled'
            elif any(i.quantity_received > 0 for i in items):
                store_req.status = 'partial'
            else:
                store_req.status = 'rejected'
        
        store_req.save()

        # Notify HOD, Store Head, and assigned staff about teacher confirmation
        teacher_name = request.user.get_full_name() or request.user.username
        status_label = dict(StoreRequest.STATUS_CHOICES).get(store_req.status, store_req.status)
        notif_msg = (
            f"Teacher '{teacher_name}' has confirmed receipt for request '{store_req.title}'. "
            f"Final Status: {status_label}."
            + (f" Teacher Remarks: {teacher_remarks}" if teacher_remarks else "")
        )
        notify_list = list(_get_store_head_users())
        if store_req.assigned_to:
            notify_list.append(store_req.assigned_to.user)
        if store_req.fulfilled_by and store_req.fulfilled_by != store_req.assigned_to:
            notify_list.append(store_req.fulfilled_by.user)
        # Notify HOD
        try:
            dept = request.user.teacher_profile.department
            hod_user = _get_dept_hod_user(dept)
            if hod_user and hod_user != request.user:
                notify_list.append(hod_user)
        except Exception:
            pass
        send_store_notification(notify_list, notif_msg, store_request=store_req, notif_type='fulfillment')

        messages.success(request, "Items receipt confirmed successfully!")
        return redirect('my_store_requests')

    return render(request, 'teacher_confirm_receipt.html', {'store_req': store_req})


# ── Teacher: track their own store request statuses ───────────────────
@login_required
def my_store_requests(request):
    """Teacher sees all their store requests, and if they are an HOD, they also see pending approvals for their dept."""
    my_requests = StoreRequest.objects.filter(
        requested_by=request.user
    ).prefetch_related('items').select_related(
        'hod_approved_by', 'assigned_to__user', 'fulfilled_by__user'
    )
    
    context = {'my_requests': my_requests, 'is_hod': False}
    
    # Check if user is HOD
    hod_profile = _get_hod_profile(request.user)
    if hod_profile:
        context['is_hod'] = True
        context['dept'] = hod_profile.department
        
        dept_teacher_ids = User.objects.filter(
            teacher_profile__department=hod_profile.department,
            is_staff=True, is_superuser=False
        ).values_list('id', flat=True)
        
        context['pending'] = StoreRequest.objects.filter(
            requested_by__in=dept_teacher_ids, status='pending_hod'
        ).prefetch_related('items').select_related('requested_by')

        context['reviewed'] = StoreRequest.objects.filter(
            requested_by__in=dept_teacher_ids
        ).exclude(status='pending_hod').prefetch_related('items').select_related(
            'requested_by', 'hod_approved_by', 'assigned_to__user', 'fulfilled_by__user'
        )
        
    return render(request, 'my_store_requests.html', context)


# ═══════════════════════════════════════════════════════════════
#  STORE NOTIFICATION API VIEWS
# ═══════════════════════════════════════════════════════════════

@login_required
def get_store_notifications(request):
    """API: Return unread store notifications for the logged-in user."""
    notifs = StoreNotification.objects.filter(
        recipient=request.user, is_read=False
    ).select_related('store_request')[:20]
    data = [{
        'id':          n.id,
        'message':     n.message,
        'type':        n.notif_type,
        'created_at':  n.created_at.strftime('%d %b %Y, %H:%M'),
        'request_id':  n.store_request_id,
        'request_title': n.store_request.title if n.store_request else '',
    } for n in notifs]
    return JsonResponse({'status': 'ok', 'notifications': data, 'count': len(data)})


@login_required
@csrf_exempt
def mark_store_notification_read(request, notif_id):
    """API: Mark a single store notification as read."""
    if request.method == 'POST':
        try:
            notif = StoreNotification.objects.get(id=notif_id, recipient=request.user)
            notif.is_read = True
            notif.save()
            return JsonResponse({'status': 'ok'})
        except StoreNotification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)
    return JsonResponse({'status': 'error'}, status=400)


@login_required
@csrf_exempt
def mark_all_store_notifications_read(request):
    """API: Mark all store notifications as read for the logged-in user."""
    if request.method == 'POST':
        StoreNotification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def submit_assignment(request, material_id):
    """Student: Submit/Resubmit assignment PDF."""
    if not hasattr(request.user, 'student'):
        messages.error(request, 'Access denied. Student profile required.')
        return redirect('home')

    student = request.user.student
    assignment = get_object_or_404(CourseMaterial, id=material_id, material_type='assignment')

    if request.method == 'POST':
        if assignment.due_date and timezone.now() > assignment.due_date:
            has_approved_request = LateSubmissionRequest.objects.filter(
                student=student,
                assignment=assignment,
                status='Approved'
            ).exists()
            if not has_approved_request:
                messages.error(request, f"Submission failed: The deadline has passed. Please submit an extension request and wait for approval.")
                return redirect('student_materials')

        file = request.FILES.get('file')
        remarks = request.POST.get('remarks', '').strip()

        if not file:
            messages.error(request, 'Please select a PDF file to upload.')
        elif not file.name.lower().endswith('.pdf'):
            messages.error(request, 'Only PDF files are allowed.')
        else:
            try:
                submission, created = StudentSubmission.objects.update_or_create(
                    student=student,
                    assignment=assignment,
                    defaults={
                        'file': file,
                        'remarks': remarks,
                        'submitted_at': timezone.now()
                    }
                )
                if created:
                    messages.success(request, f"Assignment '{assignment.title}' submitted successfully!")
                else:
                    messages.success(request, f"Assignment '{assignment.title}' re-submitted successfully!")
            except Exception as e:
                messages.error(request, f"Error submitting assignment: {e}")

    return redirect('student_materials')


@login_required
def view_submissions(request, material_id):
    """Teacher/Admin: View all student submissions for a specific assignment."""
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('home')

    assignment = get_object_or_404(CourseMaterial, id=material_id, material_type='assignment')
    
    if request.user.is_staff and not request.user.is_superuser:
        if assignment.teacher != request.user:
            messages.error(request, 'Access denied. You are not the owner of this assignment.')
            return redirect('teacher_materials')

    submissions = StudentSubmission.objects.filter(assignment=assignment).select_related('student__user')
    late_requests = LateSubmissionRequest.objects.filter(assignment=assignment).select_related('student__user')

    return render(request, 'teacher_portal/view_submissions.html', {
        'assignment': assignment,
        'submissions': submissions,
        'late_requests': late_requests,
    })


@login_required
def grade_submission(request, submission_id):
    """Teacher/Admin: Grade and provide feedback for a student submission."""
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('home')

    submission = get_object_or_404(StudentSubmission, id=submission_id)
    assignment = submission.assignment

    if request.user.is_staff and not request.user.is_superuser:
        if assignment.teacher != request.user:
            messages.error(request, 'Access denied.')
            return redirect('teacher_materials')

    if request.method == 'POST':
        grade = request.POST.get('grade', '').strip()
        feedback = request.POST.get('feedback', '').strip()

        try:
            submission.grade = grade
            submission.feedback = feedback
            submission.save()
            messages.success(request, f"Submission graded successfully for {submission.student.name}.")
        except Exception as e:
            messages.error(request, f"Error saving grade: {e}")

    return redirect('view_submissions', material_id=assignment.id)


@login_required
def request_late_submission(request, material_id):
    """Student: Submit a late submission / extension request."""
    if not hasattr(request.user, 'student'):
        messages.error(request, 'Access denied. Student profile required.')
        return redirect('home')

    student = request.user.student
    assignment = get_object_or_404(CourseMaterial, id=material_id, material_type='assignment')

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()

        if not reason:
            messages.error(request, 'Please provide a reason for your request.')
        else:
            try:
                LateSubmissionRequest.objects.update_or_create(
                    student=student,
                    assignment=assignment,
                    defaults={
                        'reason': reason,
                        'status': 'Pending',
                        'resolved_at': None
                    }
                )
                messages.success(request, f"Late submission request for '{assignment.title}' submitted successfully. Waiting for teacher approval.")
            except Exception as e:
                messages.error(request, f"Error submitting request: {e}")

    return redirect('student_materials')


@login_required
def resolve_late_request(request, request_id):
    """Teacher/Admin: Approve or Reject a student's late submission request."""
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('home')

    late_req = get_object_or_404(LateSubmissionRequest, id=request_id)
    assignment = late_req.assignment

    if request.user.is_staff and not request.user.is_superuser:
        if assignment.teacher != request.user:
            messages.error(request, 'Access denied.')
            return redirect('teacher_materials')

    if request.method == 'POST':
        action = request.POST.get('action') # 'approve' or 'reject'

        try:
            if action == 'approve':
                late_req.status = 'Approved'
                messages.success(request, f"Approved late submission request for {late_req.student.name}.")
            else:
                late_req.status = 'Rejected'
                messages.warning(request, f"Rejected late submission request for {late_req.student.name}.")
            late_req.resolved_at = timezone.now()
            late_req.save()
        except Exception as e:
            messages.error(request, f"Error updating request: {e}")

    return redirect('view_submissions', material_id=assignment.id)


@login_required
def student_applications(request):
    try:
        student = request.user.student
    except Student.DoesNotExist:
        messages.error(request, 'Access denied. Only students can view this page.')
        return redirect('index')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        attachment = request.FILES.get('attachment')

        if not title or not description:
            messages.error(request, 'Title and description are required.')
        else:
            # Resolve coordinator
            coordinator = ClassCoordinator.objects.filter(
                department=student.department,
                year=student.year,
                section=student.section
            ).first()
            
            if not coordinator:
                dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
                if student.section and student.section.upper() in dept_codes:
                    coordinator = ClassCoordinator.objects.filter(
                        year=student.year,
                        section=student.section
                    ).first()

            StudentApplication.objects.create(
                student=student,
                title=title,
                description=description,
                attachment=attachment
            )
            
            if coordinator:
                messages.success(request, f"Application '{title}' submitted successfully to coordinator {coordinator.teacher.get_full_name() or coordinator.teacher.username}.")
            else:
                messages.warning(request, f"Application '{title}' submitted, but no Class Coordinator is currently assigned for your class.")

            return redirect('student_applications')

    applications = StudentApplication.objects.filter(student=student)
    coordinator = ClassCoordinator.objects.filter(
        department=student.department,
        year=student.year,
        section=student.section
    ).first()
    
    if not coordinator:
        dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
        if student.section and student.section.upper() in dept_codes:
            coordinator = ClassCoordinator.objects.filter(
                year=student.year,
                section=student.section
            ).first()

    return render(request, 'student_portal/applications.html', {
        'student': student,
        'applications': applications,
        'coordinator': coordinator
    })


@login_required
def coordinator_dashboard(request):
    if not request.user.is_staff:
        messages.error(request, 'Access denied.')
        return redirect('index')

    coordinators = ClassCoordinator.objects.filter(teacher=request.user)
    if not coordinators.exists():
        return render(request, 'teacher_portal/coordinator_dashboard.html', {
            'no_coord': True
        })

    coord_id = request.GET.get('coordinator_id')
    if coord_id:
        current_coord = get_object_or_404(ClassCoordinator, id=coord_id, teacher=request.user)
    else:
        current_coord = coordinators.first()

    students = Student.objects.filter(
        year=current_coord.year,
        section=current_coord.section
    )
    dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
    if current_coord.section and current_coord.section.upper() in dept_codes:
        students = students.filter(department__icontains=current_coord.section)
    else:
        students = students.filter(department__icontains=current_coord.department)
    students = students.order_by('roll_number')

    applications = StudentApplication.objects.filter(student__in=students)
    pending_applications_count = applications.filter(status='Pending').count()

    # Attendance Matrix
    records = AttendanceRecord.objects.filter(student__in=students)
    attendance_cols = records.values_list('date', 'subject').distinct().order_by('-date', 'subject')[:30]

    cols_headers = [{'date': date, 'subject': subject} for date, subject in attendance_cols]

    attendance_matrix = []
    for s in students:
        s_records = { (r.date, r.subject): r.status for r in records.filter(student=s) }
        statuses = []
        for date, subject in attendance_cols:
            status = s_records.get((date, subject), '-')
            statuses.append(status)
        
        present_count = sum(1 for status in statuses if status == 'Present')
        total_count = sum(1 for status in statuses if status in ('Present', 'Absent'))
        pct = round((present_count / total_count * 100) if total_count > 0 else 0, 1)

        attendance_matrix.append({
            'student': s,
            'statuses': statuses,
            'present_count': present_count,
            'total_count': total_count,
            'percentage': pct
        })

    return render(request, 'teacher_portal/coordinator_dashboard.html', {
        'no_coord': False,
        'coordinators': coordinators,
        'current_coord': current_coord,
        'students': students,
        'applications': applications,
        'pending_applications_count': pending_applications_count,
        'cols_headers': cols_headers,
        'attendance_matrix': attendance_matrix,
    })


@login_required
def coordinator_resolve_application(request, app_id):
    if not request.user.is_staff:
        messages.error(request, 'Access denied.')
        return redirect('index')

    application = get_object_or_404(StudentApplication, id=app_id)
    student = application.student

    # Find if the logged-in teacher is the coordinator for this student
    coordinators = ClassCoordinator.objects.filter(teacher=request.user)
    is_coordinator = False
    dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
    
    for coord in coordinators:
        if coord.year == student.year and coord.section == student.section:
            if coord.section and coord.section.upper() in dept_codes:
                if coord.section.upper() in student.department.upper():
                    is_coordinator = True
                    break
            else:
                if coord.department.upper() in student.department.upper():
                    is_coordinator = True
                    break

    if not is_coordinator:
        messages.error(request, 'Access denied. You are not the coordinator for this student.')
        return redirect('coordinator_dashboard')

    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()

        if action == 'approve':
            application.status = 'Approved'
            messages.success(request, f"Application of {application.student.name} approved.")
        elif action == 'reject':
            application.status = 'Rejected'
            messages.warning(request, f"Application of {application.student.name} rejected.")
        
        application.remarks = remarks
        application.save()

    return redirect('coordinator_dashboard')


@login_required
def export_coordinator_attendance_excel(request, coord_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('index')

    if request.user.is_superuser:
        coordinator = get_object_or_404(ClassCoordinator, id=coord_id)
    else:
        coordinator = get_object_or_404(ClassCoordinator, id=coord_id, teacher=request.user)

    students = Student.objects.filter(
        year=coordinator.year,
        section=coordinator.section
    )
    dept_codes = ['CSE', 'IT', 'ECE', 'EE', 'ME', 'CE', 'AIDS', 'AIML']
    if coordinator.section and coordinator.section.upper() in dept_codes:
        students = students.filter(department__icontains=coordinator.section)
    else:
        students = students.filter(department__icontains=coordinator.department)
    students = students.order_by('roll_number')

    records = AttendanceRecord.objects.filter(student__in=students)
    col_pairs = list(records.values_list('date', 'subject').distinct().order_by('date', 'subject'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Sheet"
    ws.views.sheetView[0].showGridLines = True

    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    max_col = max(3, 4 + len(col_pairs) - 1)

    # 1. Title Row (Row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws['A1']
    YEAR_ROMAN = {1: 'I', 2: 'II', 3: 'III', 4: 'IV'}
    try:
        yr_val = int(coordinator.year)
        roman_year = YEAR_ROMAN.get(yr_val, coordinator.year)
    except ValueError:
        roman_year = coordinator.year

    title_cell.value = f"Sri Aurobindo Institute of Technology, Indore, {coordinator.department} dept. Attendance Record {roman_year} YEAR {coordinator.section}"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="5C59E8", end_color="5C59E8", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).border = thin_border

    # 2. Date Row (Row 2)
    ws['B2'] = "Date------>"
    ws['B2'].font = Font(name="Calibri", size=10, italic=True)
    ws['B2'].alignment = Alignment(horizontal="right", vertical="center")
    
    current_date = None
    start_col = 4
    
    for idx, (date, subject) in enumerate(col_pairs):
        col_idx = 4 + idx
        cell = ws.cell(row=2, column=col_idx)
        date_str = date.strftime('%d/%m/%y %a') if date else "-"
        cell.value = date_str
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=10, bold=True)
        
        if current_date is None:
            current_date = date
            start_col = col_idx
        elif date != current_date:
            if col_idx - 1 > start_col:
                ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=col_idx - 1)
            current_date = date
            start_col = col_idx

    if len(col_pairs) > 0 and (4 + len(col_pairs) - 1) > start_col:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=4 + len(col_pairs) - 1)

    # 3. Lecture Row (Row 3)
    ws['B3'] = "Lecture No ------>"
    ws['B3'].font = Font(name="Calibri", size=10, italic=True)
    ws['B3'].alignment = Alignment(horizontal="right", vertical="center")
    
    lecture_counters = {}
    for idx, (date, subject) in enumerate(col_pairs):
        col_idx = 4 + idx
        lecture_counters[date] = lecture_counters.get(date, 0) + 1
        cell = ws.cell(row=3, column=col_idx)
        cell.value = f"L{lecture_counters[date]}"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=10)

    # 4. Subject Row (Row 4)
    ws['B4'] = "SUBJECT"
    ws['B4'].font = Font(name="Calibri", size=10, bold=True)
    ws['B4'].alignment = Alignment(horizontal="left", vertical="center")
    ws['C4'] = "------->"
    ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, (date, subject) in enumerate(col_pairs):
        col_idx = 4 + idx
        cell = ws.cell(row=4, column=col_idx)
        cell.value = subject
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(name="Calibri", size=10, bold=True)

    # 5. Faculty Row (Row 5)
    ws['B5'] = "FACULTY NAME"
    ws['B5'].font = Font(name="Calibri", size=10, bold=True)
    ws['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws['C5'] = "------->"
    ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, (date, subject) in enumerate(col_pairs):
        col_idx = 4 + idx
        cell = ws.cell(row=5, column=col_idx)
        
        initials = "-"
        ts = TeacherSubject.objects.filter(
            subject__iexact=subject,
            year=coordinator.year,
            section=coordinator.section
        ).first()
        if not ts:
            ts = TeacherSubject.objects.filter(subject__iexact=subject).first()
            
        if ts and ts.teacher:
            first = ts.teacher.first_name
            last = ts.teacher.last_name
            if first and last:
                initials = f"{first[0].upper()}{last[0].upper()}"
            elif first:
                initials = first[:2].upper()
            else:
                initials = ts.teacher.username[:3].upper()
                
        cell.value = initials
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=10)

    # 6. Headers (Row 6)
    ws['A6'] = "S.No."
    ws['B6'] = "Enrollment_No"
    ws['C6'] = "Student_Name"
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for cell_ref in ['A6', 'B6', 'C6']:
        cell = ws[cell_ref]
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        
    for idx, (date, subject) in enumerate(col_pairs):
        col_idx = 4 + idx
        cell = ws.cell(row=6, column=col_idx)
        cell.value = idx + 1
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = header_fill

    for r in range(2, 7):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

    # 7. Student rows (Row 7+)
    fill_present = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    font_present = Font(name="Calibri", size=10, bold=True, color="2E7D32")
    fill_absent = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    font_absent = Font(name="Calibri", size=10, bold=True, color="C62828")
    font_none = Font(name="Calibri", size=10, color="757575")

    for s_idx, s in enumerate(students):
        row_idx = 7 + s_idx
        
        c_sno = ws.cell(row=row_idx, column=1, value=s_idx + 1)
        c_sno.alignment = Alignment(horizontal="center", vertical="center")
        
        c_roll = ws.cell(row=row_idx, column=2, value=s.roll_number)
        c_roll.alignment = Alignment(horizontal="center", vertical="center")
        
        c_name = ws.cell(row=row_idx, column=3, value=s.name)
        c_name.alignment = Alignment(horizontal="left", vertical="center")
        
        for c in [c_sno, c_roll, c_name]:
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border
            
        s_records = { (r.date, r.subject): r.status for r in records.filter(student=s) }
        
        for col_i, (date, subject) in enumerate(col_pairs):
            col_idx = 4 + col_i
            status = s_records.get((date, subject), '-')
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if status == 'Present':
                cell.value = 1
                cell.fill = fill_present
                cell.font = font_present
            elif status == 'Absent':
                cell.value = 0
                cell.fill = fill_absent
                cell.font = font_absent
            else:
                cell.value = '-'
                cell.font = font_none

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20
    
    for r in range(7, 7 + len(students)):
        ws.row_dimensions[r].height = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"attendance_{coordinator.department}_{coordinator.year}_{coordinator.section}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
